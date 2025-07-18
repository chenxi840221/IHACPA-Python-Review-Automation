#!/usr/bin/env python3
"""Analyze SNYK URL patterns in the source Excel file"""

from openpyxl import load_workbook

source_file = '02-Source-Data/2025-07-09 IHACPA Review of ALL existing PYTHON Packages.xlsx'

wb = load_workbook(source_file, read_only=True, data_only=False)
sheet = wb.active

print('=== Analyzing URL formula patterns in source file ===')

# Count different URL patterns
search_pattern = 0
pip_pattern = 0
other_pattern = 0

for row in range(4, 250):  # Check first 250 packages
    snyk_cell = sheet.cell(row=row, column=19)
    if snyk_cell.value:
        formula = str(snyk_cell.value)
        if 'vuln?search=' in formula:
            search_pattern += 1
        elif 'vuln/pip' in formula:
            pip_pattern += 1
        else:
            other_pattern += 1

print(f'\nTotal SNYK URL formulas analyzed: {search_pattern + pip_pattern + other_pattern}')
print(f'Using incorrect pattern (vuln?search=): {search_pattern}')
print(f'Using correct pattern (vuln/pip/): {pip_pattern}')
print(f'Other patterns: {other_pattern}')

# Find where the pattern changes
print('\n=== Looking for pattern changes ===')
last_pattern = None
for row in range(4, 250):
    snyk_cell = sheet.cell(row=row, column=19)
    package_name = sheet.cell(row=row, column=2).value
    
    if snyk_cell.value:
        formula = str(snyk_cell.value)
        if 'vuln?search=' in formula:
            current_pattern = 'search'
        elif 'vuln/pip' in formula:
            current_pattern = 'pip'
        else:
            current_pattern = 'other'
        
        if last_pattern and current_pattern != last_pattern:
            print(f'Pattern change at row {row} ({package_name}): {last_pattern} → {current_pattern}')
            print(f'  Formula: {formula[:100]}...')
        
        last_pattern = current_pattern

wb.close()