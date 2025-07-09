#!/usr/bin/env python3
"""
Excel Handler for IHACPA Python Package Review Automation
Handles reading and writing Excel files for the package review process
"""

import openpyxl
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
import logging


class ExcelHandler:
    """Handles Excel file operations for IHACPA package review automation"""
    
    HEADER_ROW = 3
    DATA_START_ROW = 4
    
    COLUMN_MAPPING = {
        'index': 1,                    # A: #
        'package_name': 2,             # B: Package Name
        'current_version': 3,          # C: Version
        'pypi_current_link': 4,        # D: PyPi Links (installed version)
        'date_published': 5,           # E: Date Published
        'latest_version': 6,           # F: Latest Version
        'pypi_latest_link': 7,         # G: PyPi Links (latest version)
        'latest_release_date': 8,      # H: Latest Version Release Date
        'requires': 9,                 # I: Requires
        'development_status': 10,      # J: Development Status
        'github_url': 11,              # K: GitHub URL
        'github_advisory_url': 12,     # L: GitHub Mirror Security Advisory Lookup URL
        'github_advisory_result': 13,  # M: GitHub Security Advisory Result
        'notes': 14,                   # N: Notes
        'nist_nvd_url': 15,           # O: NIST NVD Lookup URL
        'nist_nvd_result': 16,        # P: NIST NVD Lookup Result
        'mitre_cve_url': 17,          # Q: MITRE CVE Lookup URL
        'mitre_cve_result': 18,       # R: MITRE CVE Lookup Result
        'snyk_url': 19,               # S: SNYK Vulnerability Lookup URL
        'snyk_result': 20,            # T: SNYK Vulnerability Lookup Result
        'exploit_db_url': 21,         # U: Exploit Database Lookup URL
        'exploit_db_result': 22,      # V: Exploit Database Lookup Result
        'recommendation': 23          # W: Recommendation
    }
    
    def __init__(self, file_path: str):
        """Initialize Excel handler with file path"""
        self.file_path = Path(file_path)
        self.workbook = None
        self.worksheet = None
        self.logger = logging.getLogger(__name__)
        
    def load_workbook(self) -> bool:
        """Load Excel workbook and get active worksheet"""
        try:
            if not self.file_path.exists():
                self.logger.error(f"Excel file not found: {self.file_path}")
                return False
                
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            self.logger.info(f"Successfully loaded Excel file: {self.file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error loading Excel file: {e}")
            return False
    
    def get_package_count(self) -> int:
        """Get total number of packages in the Excel file"""
        if not self.worksheet:
            return 0
            
        count = 0
        for row in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
            package_name = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if package_name:
                count += 1
        return count
    
    def get_package_data(self, row_number: int) -> Dict[str, Any]:
        """Get package data for a specific row"""
        if not self.worksheet:
            return {}
            
        package_data = {}
        for field, column in self.COLUMN_MAPPING.items():
            cell_value = self.worksheet.cell(row=row_number, column=column).value
            package_data[field] = cell_value
            
        return package_data
    
    def get_all_packages(self) -> List[Dict[str, Any]]:
        """Get all package data from Excel file"""
        if not self.worksheet:
            return []
            
        packages = []
        for row in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
            package_name = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if package_name:
                package_data = self.get_package_data(row)
                package_data['row_number'] = row
                packages.append(package_data)
                
        return packages
    
    def update_package_data(self, row_number: int, updates: Dict[str, Any]) -> bool:
        """Update package data for a specific row"""
        if not self.worksheet:
            return False
            
        try:
            for field, value in updates.items():
                if field in self.COLUMN_MAPPING:
                    column = self.COLUMN_MAPPING[field]
                    
                    # Fix datetime timezone issues for Excel
                    if hasattr(value, 'tzinfo') and value.tzinfo is not None:
                        # Convert timezone-aware datetime to naive datetime
                        value = value.replace(tzinfo=None)
                    
                    self.worksheet.cell(row=row_number, column=column, value=value)
                    
            self.logger.debug(f"Updated row {row_number} with {len(updates)} fields")
            return True
            
        except Exception as e:
            self.logger.error(f"Error updating row {row_number}: {e}")
            return False
    
    def save_workbook(self, backup: bool = True) -> bool:
        """Save the Excel workbook with optional backup"""
        if not self.workbook:
            return False
            
        try:
            if backup:
                backup_path = self.file_path.with_suffix(f'.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
                self.workbook.save(backup_path)
                self.logger.info(f"Backup created: {backup_path}")
            
            self.workbook.save(self.file_path)
            self.logger.info(f"Excel file saved: {self.file_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {e}")
            return False
    
    def get_packages_by_range(self, start_row: int, end_row: int) -> List[Dict[str, Any]]:
        """Get packages within a specific row range"""
        if not self.worksheet:
            return []
            
        packages = []
        for row in range(max(start_row, self.DATA_START_ROW), min(end_row + 1, self.worksheet.max_row + 1)):
            package_name = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if package_name:
                package_data = self.get_package_data(row)
                package_data['row_number'] = row
                packages.append(package_data)
                
        return packages
    
    def find_package_by_name(self, package_name: str) -> Optional[Dict[str, Any]]:
        """Find a package by name and return its data"""
        if not self.worksheet:
            return None
            
        for row in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
            cell_value = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if cell_value and str(cell_value).strip().lower() == package_name.strip().lower():
                package_data = self.get_package_data(row)
                package_data['row_number'] = row
                return package_data
                
        return None
    
    def get_packages_needing_update(self) -> List[Dict[str, Any]]:
        """Get packages that need automated updates (missing data in automated columns)"""
        if not self.worksheet:
            return []
            
        automated_fields = [
            'date_published', 'latest_version', 'pypi_latest_link', 'latest_release_date',
            'requires', 'development_status', 'github_url', 'github_advisory_url',
            'github_advisory_result', 'nist_nvd_url', 'nist_nvd_result',
            'mitre_cve_url', 'mitre_cve_result', 'snyk_url', 'snyk_result',
            'exploit_db_url', 'exploit_db_result', 'recommendation'
        ]
        
        packages_needing_update = []
        for row in range(self.DATA_START_ROW, self.worksheet.max_row + 1):
            package_name = self.worksheet.cell(row=row, column=self.COLUMN_MAPPING['package_name']).value
            if package_name:
                package_data = self.get_package_data(row)
                package_data['row_number'] = row
                
                # Check if any automated field is empty
                needs_update = False
                for field in automated_fields:
                    if not package_data.get(field):
                        needs_update = True
                        break
                
                if needs_update:
                    packages_needing_update.append(package_data)
                    
        return packages_needing_update
    
    def get_file_info(self) -> Dict[str, Any]:
        """Get general information about the Excel file"""
        if not self.worksheet:
            return {}
            
        return {
            'file_path': str(self.file_path),
            'sheet_names': self.workbook.sheetnames,
            'total_rows': self.worksheet.max_row,
            'total_columns': self.worksheet.max_column,
            'header_row': self.HEADER_ROW,
            'data_start_row': self.DATA_START_ROW,
            'package_count': self.get_package_count(),
            'last_modified': datetime.fromtimestamp(self.file_path.stat().st_mtime) if self.file_path.exists() else None
        }
    
    def validate_file_structure(self) -> Tuple[bool, List[str]]:
        """Validate that the Excel file has the expected structure"""
        if not self.worksheet:
            return False, ["Excel file not loaded"]
            
        errors = []
        
        # Check if we have the expected number of columns
        if self.worksheet.max_column < max(self.COLUMN_MAPPING.values()):
            errors.append(f"Expected at least {max(self.COLUMN_MAPPING.values())} columns, found {self.worksheet.max_column}")
        
        # Check header row for expected column names
        expected_headers = [
            "#", "Package Name", "Version", "PyPi Links", "Date Published",
            "Latest Version", "PyPi Links", "Latest Version Release Date",
            "Requires", "Development Status", "GitHub URL"
        ]
        
        for i, expected_header in enumerate(expected_headers[:5], 1):  # Check first 5 headers
            actual_header = self.worksheet.cell(row=self.HEADER_ROW, column=i).value
            if not actual_header or expected_header.lower() not in str(actual_header).lower():
                errors.append(f"Column {i} header mismatch. Expected '{expected_header}', found '{actual_header}'")
        
        # Check if we have package data
        if self.get_package_count() == 0:
            errors.append("No package data found in Excel file")
        
        return len(errors) == 0, errors
    
    def compare_with_original(self, original_file_path: str) -> Dict[str, Any]:
        """Compare current Excel state with original file"""
        comparison_results = {
            'total_changes': 0,
            'packages_modified': 0,
            'changes_by_package': {},
            'changes_by_column': {},
            'summary': []
        }
        
        try:
            # Load original file
            original_workbook = openpyxl.load_workbook(original_file_path)
            original_worksheet = original_workbook.active
            
            # Get all packages from current file
            current_packages = self.get_all_packages()
            
            for package in current_packages:
                package_name = package.get('package_name', '')
                row_number = package.get('row_number', 0)
                
                if not package_name or not row_number:
                    continue
                
                changes = []
                
                # Compare each field
                for field, column in self.COLUMN_MAPPING.items():
                    current_value = package.get(field)
                    original_value = original_worksheet.cell(row=row_number, column=column).value
                    
                    # Handle datetime comparison
                    if hasattr(current_value, 'tzinfo') and current_value.tzinfo is not None:
                        current_value = current_value.replace(tzinfo=None)
                    
                    # Compare values (handle None vs empty string)
                    current_str = str(current_value) if current_value is not None else ""
                    original_str = str(original_value) if original_value is not None else ""
                    
                    if current_str != original_str and current_str != "":
                        changes.append({
                            'field': field,
                            'column': self._get_column_letter(column),
                            'original_value': original_str,
                            'new_value': current_str,
                            'change_type': 'updated' if original_str else 'added'
                        })
                        
                        # Track changes by column
                        col_letter = self._get_column_letter(column)
                        if col_letter not in comparison_results['changes_by_column']:
                            comparison_results['changes_by_column'][col_letter] = 0
                        comparison_results['changes_by_column'][col_letter] += 1
                
                if changes:
                    comparison_results['packages_modified'] += 1
                    comparison_results['changes_by_package'][package_name] = {
                        'row_number': row_number,
                        'changes': changes,
                        'change_count': len(changes)
                    }
                    comparison_results['total_changes'] += len(changes)
            
            # Generate summary
            comparison_results['summary'] = [
                f"Total packages modified: {comparison_results['packages_modified']}",
                f"Total field changes: {comparison_results['total_changes']}",
                f"Most changed columns: {self._get_top_changed_columns(comparison_results['changes_by_column'])}"
            ]
            
            original_workbook.close()
            
        except Exception as e:
            self.logger.error(f"Error comparing with original file: {e}")
            comparison_results['error'] = str(e)
        
        return comparison_results
    
    def _get_column_letter(self, column_number: int) -> str:
        """Convert column number to Excel column letter"""
        from openpyxl.utils import get_column_letter
        return get_column_letter(column_number)
    
    def _get_top_changed_columns(self, changes_by_column: Dict[str, int]) -> str:
        """Get the top 5 most changed columns"""
        if not changes_by_column:
            return "None"
        
        sorted_columns = sorted(changes_by_column.items(), key=lambda x: x[1], reverse=True)
        top_5 = sorted_columns[:5]
        return ", ".join([f"{col} ({count})" for col, count in top_5])
    
    def generate_changes_report(self, comparison_results: Dict[str, Any]) -> str:
        """Generate a detailed changes report"""
        report = []
        report.append("EXCEL FILE CHANGES REPORT")
        report.append("=" * 50)
        report.append("")
        
        # Summary
        for summary_line in comparison_results.get('summary', []):
            report.append(summary_line)
        report.append("")
        
        # Detailed changes by package
        if comparison_results['packages_modified'] > 0:
            report.append("DETAILED CHANGES BY PACKAGE:")
            report.append("-" * 30)
            
            for package_name, package_data in comparison_results['changes_by_package'].items():
                report.append(f"\n📦 {package_name} (Row {package_data['row_number']}):")
                
                for change in package_data['changes']:
                    field_name = change['field'].replace('_', ' ').title()
                    if change['change_type'] == 'added':
                        report.append(f"  ✅ {field_name} (Col {change['column']}): Added '{change['new_value']}'")
                    else:
                        report.append(f"  🔄 {field_name} (Col {change['column']}): '{change['original_value']}' → '{change['new_value']}'")
        
        # Column summary
        if comparison_results['changes_by_column']:
            report.append("\nCHANGES BY COLUMN:")
            report.append("-" * 20)
            
            for col, count in sorted(comparison_results['changes_by_column'].items()):
                report.append(f"Column {col}: {count} changes")
        
        return "\n".join(report)
    
    def close(self):
        """Close the Excel workbook"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.worksheet = None